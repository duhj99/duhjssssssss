import os
import re
import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

class ExcelProcessor:
    @staticmethod
    def find_replace(file_path, find_text, replace_text, sheet_range=None, use_regex=False, output_dir=None):
        """
        在Excel文件中查找并替换文本
        
        Args:
            file_path: Excel文件路径
            find_text: 要查找的文本
            replace_text: 替换的文本
            sheet_range: 工作表范围，格式为"Sheet1!A1:C10"，如果为None则处理所有工作表
            use_regex: 是否使用正则表达式
            output_dir: 输出目录，如果为None则覆盖原文件
            
        Returns:
            处理后的文件路径
        """
        try:
            # 创建输出文件路径
            if output_dir:
                output_path = os.path.join(output_dir, os.path.basename(file_path))
            else:
                output_path = file_path
                
            # 如果输出路径与输入路径不同，先复制文件
            if output_path != file_path:
                wb = openpyxl.load_workbook(file_path)
                wb.save(output_path)
            else:
                wb = openpyxl.load_workbook(file_path)
            
            # 解析工作表范围
            if sheet_range:
                parts = sheet_range.split('!')
                sheet_name = parts[0]
                cell_range = parts[1] if len(parts) > 1 else None
                
                if sheet_name not in wb.sheetnames:
                    raise ValueError(f"工作表 '{sheet_name}' 不存在")
                
                sheets = [wb[sheet_name]]
            else:
                sheets = wb.worksheets
                cell_range = None
            
            # 处理每个工作表
            for sheet in sheets:
                # 如果指定了单元格范围
                if cell_range:
                    cells = sheet[cell_range]
                    # 如果是单个单元格
                    if not isinstance(cells, tuple):
                        cells = [[cells]]
                    
                    for row in cells:
                        for cell in row:
                            if cell.value and isinstance(cell.value, str):
                                if use_regex:
                                    cell.value = re.sub(find_text, replace_text, cell.value)
                                else:
                                    cell.value = cell.value.replace(find_text, replace_text)
                else:
                    # 处理整个工作表
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value and isinstance(cell.value, str):
                                if use_regex:
                                    cell.value = re.sub(find_text, replace_text, cell.value)
                                else:
                                    cell.value = cell.value.replace(find_text, replace_text)
            
            # 保存工作簿
            wb.save(output_path)
            return output_path
        except Exception as e:
            raise Exception(f"处理Excel文件时出错: {str(e)}")
    
    @staticmethod
    def batch_find_replace(file_paths, replacements, sheet_range=None, use_regex=False, output_dir=None):
        """
        批量处理多个Excel文件的查找替换
        
        Args:
            file_paths: Excel文件路径列表
            replacements: 替换规则列表，每个规则是一个(find_text, replace_text)元组
            sheet_range: 工作表范围，格式为"Sheet1!A1:C10"，如果为None则处理所有工作表
            use_regex: 是否使用正则表达式
            output_dir: 输出目录
            
        Returns:
            处理后的文件路径列表
        """
        processed_files = []
        
        for file_path in file_paths:
            # 创建输出文件路径
            if output_dir:
                output_path = os.path.join(output_dir, os.path.basename(file_path))
            else:
                output_path = file_path
                
            # 如果输出路径与输入路径不同，先复制文件
            if output_path != file_path:
                wb = openpyxl.load_workbook(file_path)
                wb.save(output_path)
            else:
                wb = openpyxl.load_workbook(file_path)
            
            # 解析工作表范围
            if sheet_range:
                parts = sheet_range.split('!')
                sheet_name = parts[0]
                cell_range = parts[1] if len(parts) > 1 else None
                
                if sheet_name not in wb.sheetnames:
                    raise ValueError(f"工作表 '{sheet_name}' 不存在")
                
                sheets = [wb[sheet_name]]
            else:
                sheets = wb.worksheets
                cell_range = None
            
            # 应用所有替换规则
            for find_text, replace_text in replacements:
                # 处理每个工作表
                for sheet in sheets:
                    # 如果指定了单元格范围
                    if cell_range:
                        cells = sheet[cell_range]
                        # 如果是单个单元格
                        if not isinstance(cells, tuple):
                            cells = [[cells]]
                        
                        for row in cells:
                            for cell in row:
                                if cell.value and isinstance(cell.value, str):
                                    if use_regex:
                                        cell.value = re.sub(find_text, replace_text, cell.value)
                                    else:
                                        cell.value = cell.value.replace(find_text, replace_text)
                    else:
                        # 处理整个工作表
                        for row in sheet.iter_rows():
                            for cell in row:
                                if cell.value and isinstance(cell.value, str):
                                    if use_regex:
                                        cell.value = re.sub(find_text, replace_text, cell.value)
                                    else:
                                        cell.value = cell.value.replace(find_text, replace_text)
            
            # 保存工作簿
            wb.save(output_path)
            processed_files.append(output_path)
        
        return processed_files
    
    @staticmethod
    def merge_excel_files(file_paths, merge_type="rows", output_path=None, remove_duplicates=False):
        """
        合并多个Excel文件
        
        Args:
            file_paths: Excel文件路径列表
            merge_type: 合并类型，可以是"rows"（按行合并）、"columns"（按列合并）或"sheets"（按工作表合并）
            output_path: 输出文件路径
            remove_duplicates: 是否删除重复项
            
        Returns:
            合并后的文件路径
        """
        try:
            if not file_paths:
                raise ValueError("没有提供要合并的文件")
            
            if merge_type == "sheets":
                # 按工作表合并（将每个文件作为新的工作表添加到一个工作簿中）
                wb = openpyxl.Workbook()
                # 删除默认创建的工作表
                wb.remove(wb.active)
                
                for file_path in file_paths:
                    source_wb = openpyxl.load_workbook(file_path)
                    file_name = os.path.basename(file_path)
                    
                    for sheet_name in source_wb.sheetnames:
                        source_sheet = source_wb[sheet_name]
                        # 创建新的工作表名称（文件名+工作表名）
                        new_sheet_name = f"{file_name}_{sheet_name}"
                        # 如果名称太长，截断它
                        if len(new_sheet_name) > 31:  # Excel工作表名称最大长度为31
                            new_sheet_name = new_sheet_name[:31]
                        
                        # 创建新工作表
                        new_sheet = wb.create_sheet(title=new_sheet_name)
                        
                        # 复制单元格内容
                        for row in source_sheet.iter_rows():
                            for cell in row:
                                new_sheet[cell.coordinate].value = cell.value
                
                wb.save(output_path)
                return output_path
            else:
                # 按行或列合并（使用pandas）
                dfs = []
                
                for file_path in file_paths:
                    # 读取所有工作表
                    excel_file = pd.ExcelFile(file_path)
                    for sheet_name in excel_file.sheet_names:
                        df = pd.read_excel(file_path, sheet_name=sheet_name)
                        dfs.append(df)
                
                if merge_type == "rows":
                    # 按行合并（垂直堆叠）
                    merged_df = pd.concat(dfs, ignore_index=True)
                else:  # merge_type == "columns"
                    # 按列合并（水平堆叠）
                    merged_df = pd.concat(dfs, axis=1)
                
                # 删除重复项（如果需要）
                if remove_duplicates:
                    merged_df = merged_df.drop_duplicates()
                
                # 保存到Excel文件
                merged_df.to_excel(output_path, index=False)
                return output_path
        except Exception as e:
            raise Exception(f"合并Excel文件时出错: {str(e)}")